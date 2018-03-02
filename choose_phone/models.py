from django.db import models
import openpyxl


class Phone(models.Model):
    number = models.CharField(unique=True, max_length=16)
    reserved = models.BooleanField(default=False)

    class Meta:
        verbose_name = "Номер телефона"
        verbose_name_plural = "Номера телефонов"

    def __str__(self):
        return "({}{}{}) {}{}{}-{}{}{}{}".format(*tuple(self.number))


class PhoneNumbersFile(models.Model):
    file_path = models.FileField('Выберите файл')

    @staticmethod
    def get_phones(file):
        wb_phones = openpyxl.load_workbook(file)
        worksheet = wb_phones[wb_phones.sheetnames[0]]  # берём первый лист (на нём данные)
        max_row = worksheet.max_row
        data_frame = worksheet['A1':'B{}'.format(max_row)]
        phones_list = [(str(phone.value), str(status.value)) for phone, status in data_frame]
        return phones_list

    def __str__(self):
        return self.file_path.path

    def save(self, *args):
        phones = self.get_phones(self.file_path)
        for row in phones:
            phone = Phone()
            if row[0].isdigit() and (len(row[0]) == 10):  # проверка длины строки и содержания букв
                phone.number = row[0]
                if row[1] == 'Занят':
                    phone.reserved = True
                phone.save()
        super(PhoneNumbersFile, self).save()

    class Meta:
        verbose_name = "Загруженный файл"
        verbose_name_plural = "Загруженные файлы с номерами"
