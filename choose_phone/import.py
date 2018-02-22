import openpyxl
from choose_phone.models import Phone


def get_phones(file):
    wb_phones = openpyxl.load_workbook(file)
    worksheet = wb_phones[wb_phones.sheetnames[0]]  # берём первый лист (на нём данные)
    max_row = worksheet.max_row
    data_frame = worksheet['A1':'B{}'.format(max_row)]
    phones_list = [(phone.value, status.value) for phone, status in data_frame]
    return phones_list


source_file = './Телефоны.xlsx'
phones = get_phones(source_file)
for row in phones:
    phone = Phone()
    phone.number = row[0]
    if row[1] == 'Занят':
        phone.reserve = True
    phone.save()